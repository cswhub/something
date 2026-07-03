from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from typing import Any, Optional
import time
import re
from io import BytesIO, IOBase


class ExcelManage:
    """
data = [
        ("id", "姓名", "年龄", "任务比例比例比例比例比例比例比例比例比例"),
        (1, "张三", 20, 0.4561456456489),
        (2, "李四", 21, 0.18494494949),
        (3, "王五", 18, 0.49749849149),
        (4, "赵六", 20, 1),
    ]
colwidthlist = {
            'A': 15,
            'B': 15,
            'D': 15,
        }
with ExcelManage("test.xlsx") as excel:
    excel .write_row(data)
    excel.sheet("test2").emptysheet().write_row(data).default_style_has_header().set_style_col_ratio(("D",)).set_style_col_width()
    # 上述代码中set_style_col_width()缺省为自动计算宽度，也可以指定宽度，格式为colwidthlist，可以全部指定，也可以缺省不需要指定的比如例子中的C列
    # set_style_col_ratio(("D",))为设置某列float类型的用%显示，把需要设置的列放在tuple中传入即可
    # 如果default_style_xxx无法满足需求，可以使用set_style自己设置。
    # 读取就更简单了
    exceldata = excel.read_ws()
    print("读取的结果为：")
    print(exceldata)
    # 只读取的话可以在实例化excel对象时，指定read_only参数提高效率。
    # 如果发现读取的内容是公式，而自己想要的是计算结果，请使用data_only参数。
    """

    # EXCEL_COLUMNS = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
    def __init__(self, file, read_only: bool = False, data_only: bool = False, auto_save: bool = False):
        self.ws = None
        self.temp_sheet_name = None
        self.file_path = None
        self._mem_file: BytesIO | None = None  # 内存文件句柄
        self.auto_save = auto_save
        if file is None:
            self._mem_file = BytesIO()
            wb_new = Workbook()
            wb_new.save(self._mem_file)
            self._mem_file.seek(0)
            self.wb = load_workbook(self._mem_file, read_only=read_only, data_only=data_only)
        elif isinstance(file, IOBase):
            self.wb = load_workbook(file, read_only=read_only, data_only=data_only)
        else:
            self.file_path = file
            try:
                self.wb = load_workbook(file, read_only=read_only, data_only=data_only)
            except FileNotFoundError:
                self.wb = Workbook()
                self.temp_sheet_name = f"__temp_{str(time.time())}"
                if "Sheet" in self.wb.sheetnames:
                    self.wb["Sheet"].title = self.temp_sheet_name
                else:
                    self.wb.create_sheet(self.temp_sheet_name)

        self.style_default_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        self.style_default_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        self.style_default_title_alignment = Alignment(horizontal="center", vertical="center")
        self.style_default_font = Font(name="宋体", size=9)
        self.style_default_title_font = Font(name="宋体", size=9, bold=True)

    def __enter__(self):
        return self

    def __exit__(self, exc_type: Optional[type], exc_val: Optional[Exception], exc_tb: Optional[Any]) -> None:
        if self.auto_save:
            self.save()
        self.close()

    def sheet(self, sheet_name=None):
        if sheet_name is None:
            self.ws = self.wb.active
            return self
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(sheet_name)
        self.ws = self.wb[sheet_name]
        return self

    def emptysheet(self):
        assert self.ws is not None
        temp_title = self.ws.title
        self.wb.remove(self.ws)
        self.wb.create_sheet(temp_title)
        self.ws = self.wb[temp_title]
        return self

    def read_ws(self, start_row=1, start_col=1, end_row=None, end_col=None):
        assert self.ws is not None
        sheet = self.ws
        if end_row is None:
            end_row = sheet.max_row
        if end_col is None:
            end_col = sheet.max_column

        data = []
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col,
                                   values_only=True):
            data.append(tuple(x.strip() if isinstance(x, str) else x for x in row))
        return data

    def write_row(self, ws_data):
        assert self.ws is not None
        for row in ws_data:
            self.ws.append(row)
        return self

    def set_style(self, start_row=1, start_col=1, end_row=None, end_col=None, style_font=None, style_alignment=None,
                  style_border=None, line_height=None, style_fill=None):
        assert self.ws is not None
        sheet = self.ws
        if end_row is None:
            end_row = sheet.max_row
        if end_col is None:
            end_col = sheet.max_column
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):

            for cell in row:
                if style_font:
                    cell.font = style_font
                if style_alignment:
                    cell.alignment = style_alignment
                if style_border:
                    cell.border = style_border
                if style_fill:
                    cell.fill = style_fill

        if line_height:
            for i in range(start_row, end_row + 1):
                sheet.row_dimensions[i].height = line_height
        return self

    def default_style_has_header(self):
        self.set_style(
            end_row=1,
            style_font=self.style_default_title_font,
            style_alignment=self.style_default_title_alignment,
            style_border=self.style_default_border,
            line_height=21
        )
        self.set_style(
            start_row=2,
            style_font=self.style_default_font,
            style_alignment=self.style_default_alignment,
            style_border=self.style_default_border,
        )
        return self

    def default_style_not_has_header(self):
        self.set_style(
            style_font=self.style_default_font,
            style_alignment=self.style_default_alignment,
            style_border=self.style_default_border,
        )
        return self

    def set_style_col_ratio(self, ratio_style_tuple: tuple):
        assert self.ws is not None
        sheet = self.ws
        for col in ratio_style_tuple:
            for cell in sheet[col]:
                cell.number_format = "0.00%"
        return self

    def set_style_col_datetime(self, datetime_style_tuple, datetime_pattern):
        # TODO 未完成
        assert self.ws is not None
        sheet = self.ws
        for col in datetime_style_tuple:
            for cell in sheet[col]:
                cell.number_format = datetime_pattern

    def set_style_col_width(self, col_width_dict=None):
        assert self.ws is not None
        sheet = self.ws
        if not col_width_dict:
            print("自动设置列宽")
            col_width_dict = {}
            for column in sheet.columns:
                max_length = 0
                if column[0].column is None:
                    continue
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is None:
                            continue
                        # cell_length = len(str(cell.value))
                        cell_length = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception as e:
                        print(e)

                adjusted_width = (max_length + 5)
                col_width_dict[column_letter] = adjusted_width
        for key in col_width_dict:
            sheet.column_dimensions[key].width = col_width_dict[key]
        return self

    def save(self, save_path=None):
        if self.temp_sheet_name:
            if len(self.wb.sheetnames) >= 2:
                if self.temp_sheet_name in self.wb.sheetnames:
                    del self.wb[self.temp_sheet_name]
            else:
                if self.temp_sheet_name in self.wb.sheetnames:
                    self.wb[self.temp_sheet_name].title = "Sheet"

        if save_path:
            self.file_path = save_path
        if not self.file_path:
            raise ValueError("请指定保存路径")
        self.wb.save(self.file_path)
        print(f"文件已保存到{self.file_path}")
        return self

    def close(self):
        self.wb.close()

    def to_bytes(self) -> bytes:
        """内存模式专用：返回 excel 二进制"""
        if not self._mem_file:
            raise RuntimeError("只有内存模式才能 to_bytes()")
        self._mem_file.seek(0)
        self.wb.save(self._mem_file)
        self._mem_file.seek(0)
        return self._mem_file.read()


if __name__ == '__main__':
    data = [
        ("id", "姓名", "年龄", "任务比例比例比例比例比例比例比例比例比例"),
        (1, "张三", 20, 0.4561456456489),
        (2, "李四", 21, 0.18494494949),
        (3, "王五", 18, 0.49749849149),
        (4, "赵六", 20, 1),
    ]
    colwidthlist = {
        'A': 15,
        'B': 15,
        'D': 15,
    }
    """
    with ExcelManage("test.xlsx", auto_save=False) as excel:
        excel.sheet()
        print(excel.wb.sheetnames)
        excel.write_row(data)
        excel.sheet("test1").emptysheet().write_row(data).default_style_has_header().set_style_col_ratio(
            ("D",)).set_style_col_width()
        # 上述代码中set_style_col_width()缺省为自动计算宽度，也可以指定宽度，格式为colwidthlist，可以全部指定，也可以缺省不需要指定的比如例子中的C列
        # set_style_col_ratio(("D",))为设置某列float类型的用%显示，把需要设置的列放在tuple中传入即可
        # 如果default_style_xxx无法满足需求，可以使用set_style自己设置。
        # 读取就更简单了
        exceldata = excel.read_ws()
        print("读取的结果为：")
        print(exceldata)

        # 只读取的话可以在实例化excel对象时，指定read_only参数提高效率。
        # 如果发现读取的内容是公式，而自己想要的是计算结果，请使用data_only参数。
    """
    with ExcelManage(None, auto_save=False) as excel:
        excel.sheet()
        print(excel.wb.sheetnames)
        excel.write_row(data)
        excel.sheet("test1").emptysheet().write_row(data).default_style_has_header().set_style_col_ratio(
            ("D",)).set_style_col_width()
        # 上述代码中set_style_col_width()缺省为自动计算宽度，也可以指定宽度，格式为colwidthlist，可以全部指定，也可以缺省不需要指定的比如例子中的C列
        # set_style_col_ratio(("D",))为设置某列float类型的用%显示，把需要设置的列放在tuple中传入即可
        # 如果default_style_xxx无法满足需求，可以使用set_style自己设置。
        # 读取就更简单了
        exceldata = excel.read_ws()
        print("读取的结果为：")
        print(exceldata)
        excel_bytes = excel.to_bytes()
        with open("demo_memory.xlsx", "wb") as f:
            f.write(excel_bytes)
